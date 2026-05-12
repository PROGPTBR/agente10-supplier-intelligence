"""Parse the IBGE CNAE 2.3 "Estrutura Detalhada" XLSX into taxonomy.json.

Usage:
    cd backend && uv run python scripts/parse_ibge_xls.py

Reads:  data/cnae_2.3/raw/CNAE_Subclasses_2_3_Estrutura_Detalhada.xlsx
Writes: data/cnae_2.3/taxonomy.json (1331 entries; raises if count differs)

The IBGE XLSX has six columns: A=Seção, B=Divisão, C=Grupo, D=Classe, E=Subclasse,
F=Denominação. Hierarchy levels use merged cells (so higher levels appear in their own
column, not cascaded into column A). Subclass rows are detected by a 7-digit-ish code
in column E. This workbook does not include the "Esta subclasse compreende:" notes text,
so notas_explicativas and exemplos_atividades are null for all entries.
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

RAW_PATH = (
    Path(__file__).parent.parent / "data" / "cnae_2.3" / "raw"
    / "CNAE_Subclasses_2_3_Estrutura_Detalhada.xlsx"
)
OUT_PATH = Path(__file__).parent.parent / "data" / "cnae_2.3" / "taxonomy.json"
EXPECTED_COUNT = 1331

# IBGE subclass codes are 7-digit strings; XLSX often shows them as "XXXX-X/XX" or "XX.XX-X-XX".
# We normalize to 7 digits only.
SUBCLASS_RE = re.compile(r"^\s*(\d{2})[\.\-]?(\d{2})[\-\.]?(\d)[\-\./]?(\d{2})\s*$")
DIVISAO_RE = re.compile(r"^\s*(\d{2})\s*$")
GRUPO_RE = re.compile(r"^\s*(\d{2})[\.\-](\d)\s*$")
CLASSE_RE = re.compile(r"^\s*(\d{2})[\.\-](\d{2})[\-](\d)\s*$")
SECAO_RE = re.compile(r"^\s*([A-U])\s*$")

# Column indices (0-based) in the IBGE "Estrutura Detalhada" XLSX
COL_SECAO = 0
COL_DIVISAO = 1
COL_GRUPO = 2
COL_CLASSE = 3
COL_SUBCLASSE = 4
COL_DENOM = 5


def _norm_code(raw: str) -> str | None:
    """Normalize a subclass cell to its 7-digit code, or return None if not a subclass."""
    if raw is None:
        return None
    m = SUBCLASS_RE.match(str(raw))
    if not m:
        return None
    return "".join(m.groups())


def _extract_rows(xls_path: Path) -> list[dict]:
    """Walk the workbook and emit one dict per CNAE subclass.

    Column layout (0-indexed):
      0=Seção  1=Divisão  2=Grupo  3=Classe  4=Subclasse  5=Denominação

    Higher-level columns use merged cells, so a non-empty value in col 0 means a new
    Seção, col 1 a new Divisão, etc. We forward-fill each level independently.
    A row with a non-empty col 4 that matches the subclass regex is a subclass row.
    """
    wb = load_workbook(xls_path, data_only=True)
    ws = wb.active

    current_secao: str | None = None
    current_divisao: str | None = None
    current_grupo: str | None = None
    current_classe: str | None = None
    rows: list[dict] = []

    def _cell(row: tuple, idx: int) -> str:
        v = row[idx] if idx < len(row) else None
        return str(v).strip() if v is not None else ""

    for raw_row in ws.iter_rows(values_only=True):
        secao_val = _cell(raw_row, COL_SECAO)
        divisao_val = _cell(raw_row, COL_DIVISAO)
        grupo_val = _cell(raw_row, COL_GRUPO)
        classe_val = _cell(raw_row, COL_CLASSE)
        subclasse_val = _cell(raw_row, COL_SUBCLASSE)

        # Update hierarchy cursors
        if SECAO_RE.match(secao_val):
            current_secao = secao_val
        if DIVISAO_RE.match(divisao_val):
            current_divisao = divisao_val
        if GRUPO_RE.match(grupo_val):
            current_grupo = grupo_val.replace("-", ".").replace("/", ".").replace(".", "")[:3]
        if CLASSE_RE.match(classe_val):
            current_classe = re.sub(r"\D", "", classe_val)[:5]

        # Detect subclass row
        code = _norm_code(subclasse_val)
        if code is not None:
            denominacao = _cell(raw_row, COL_DENOM)
            rows.append(
                {
                    "codigo": code,
                    "secao": current_secao,
                    "divisao": current_divisao,
                    "grupo": current_grupo,
                    "classe": current_classe,
                    "denominacao": denominacao,
                    "notas_explicativas": None,
                    "exemplos_atividades": None,
                }
            )

    return rows


def _extract_compreende(notes: str | None) -> str | None:
    """Extract only the bullets under 'Esta subclasse compreende:' / 'compreende também:'
    (stops at 'não compreende' or end). Returns None if the section isn't present.

    Strips header lines that may leak into the captured text when the IBGE notes have
    nested 'Esta subclasse também compreende:' subsections.
    """
    if not notes:
        return None
    pattern = re.compile(
        r"esta subclasse compreende[^:\n]*:(.*?)"
        r"(?=esta subclasse não compreende|notas? explicativas?|$)",
        re.IGNORECASE | re.DOTALL,
    )
    header_line = re.compile(r"\s*esta subclasse[^\n]*", re.IGNORECASE)
    cleaned: list[str] = []
    for m in pattern.finditer(notes):
        captured = m.group(1)
        # Drop any "Esta subclasse ..." header line that leaked through
        body = "\n".join(
            line for line in captured.split("\n") if not header_line.match(line)
        )
        cleaned.append(body.strip())
    if not cleaned:
        return None
    out = "\n".join(c for c in cleaned if c).strip()
    return out or None


def main() -> int:
    if not RAW_PATH.exists():
        print(
            f"ERROR: {RAW_PATH} not found.\n"
            "  Download the IBGE CNAE 2.3 Estrutura Detalhada XLSX from concla.ibge.gov.br\n"
            "  and place it at that exact path.",
            file=sys.stderr,
        )
        return 2

    rows = _extract_rows(RAW_PATH)
    print(f"Parsed {len(rows)} CNAE subclasses from {RAW_PATH.name}")
    if len(rows) != EXPECTED_COUNT:
        print(
            f"ERROR: expected {EXPECTED_COUNT} subclasses, got {len(rows)}. "
            "Check XLSX structure / parser regex.",
            file=sys.stderr,
        )
        return 1

    bad = [r for r in rows if not r["codigo"] or not r["denominacao"]]
    if bad:
        print(f"ERROR: {len(bad)} rows missing codigo/denominacao", file=sys.stderr)
        return 1

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUT_PATH.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(rows)} entries to {OUT_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
