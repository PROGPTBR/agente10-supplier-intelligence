"""Seed spend_classification_cache from a curated XLSX (e.g., ELETROBRAS pilot 'Sugestão' tab).

The Sugestão tab has merged-cell groups: AGRUPAMENTO / NOVA CLASSIFICAÇÃO / CNAE
are filled only on the first row of each group; subsequent rows in the same group
have those columns blank. We forward-fill before extracting (description, cnae) pairs.

Inserts cache entries with metodo='golden' and confianca=0.95 — high priority, won't
be downgraded by later curator/retrieval runs (see classification_cache._METHOD_PRIORITY).

Usage:
    DATABASE_URL=... uv run python scripts/seed_classification_cache.py path/to/file.xlsx \\
        --sheet 'Sugestão'
"""

from __future__ import annotations

import argparse
import asyncio
import os
import re
import sys
from pathlib import Path

import asyncpg
from openpyxl import load_workbook

# Add backend/src to path so we can reuse the cache helpers
sys.path.insert(0, str(Path(__file__).parent.parent / "src"))
from agente10.cache.classification_cache import hash_description, normalize_description  # noqa: E402

CNAE_DASHED_RE = re.compile(r"^\s*(\d{4})-?(\d)/?(\d{2})\s*$")


def _clean_cnae(raw: str | None) -> str | None:
    """Convert '7739-0/99' or '7739099' → '7739099'. Returns None if invalid."""
    if not raw:
        return None
    m = CNAE_DASHED_RE.match(str(raw).strip())
    if not m:
        return None
    return f"{m.group(1)}{m.group(2)}{m.group(3)}"


def extract_pairs(
    xlsx_path: Path, sheet_name: str
) -> list[tuple[str, str]]:
    """Walk the sheet with forward-fill and yield (descricao, cnae_7digit) pairs."""
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet {sheet_name!r} not found. Available: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(h).strip() if h else "" for h in rows[0]]
    # Locate columns by header name (case-insensitive)
    def find_col(name: str) -> int | None:
        for i, h in enumerate(header):
            if h.lower().replace(" ", "") == name.lower().replace(" ", ""):
                return i
        return None

    col_old = find_col("antiga classificação") or find_col("ANTIGA CLASSIFICAÇÃO")
    col_new = find_col("nova classificação") or find_col("NOVA CLASSIFICAÇÃO")
    col_cnae = find_col("cnae")
    if col_old is None or col_cnae is None:
        raise SystemExit(
            f"Required columns not found. Header was: {header}\n"
            f"Need: 'ANTIGA CLASSIFICAÇÃO' and 'CNAE'"
        )

    pairs: list[tuple[str, str]] = []
    seen_keys: set[str] = set()  # by normalized description
    current_new: str | None = None
    current_cnae: str | None = None

    for row in rows[1:]:
        if not row or all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        old = (str(row[col_old]).strip() if col_old is not None and row[col_old] else "") or ""
        new_raw = (str(row[col_new]).strip() if col_new is not None and row[col_new] else "") or ""
        cnae_raw = (str(row[col_cnae]).strip() if row[col_cnae] else "") or ""

        # Forward-fill group-level columns
        cnae_clean = _clean_cnae(cnae_raw)
        if cnae_clean:
            current_cnae = cnae_clean
        if new_raw:
            current_new = new_raw

        if not current_cnae:
            continue  # group has no CNAE assigned yet — skip

        # 1) original item description (ANTIGA) → current_cnae
        if old:
            key = normalize_description(old)
            if key and key not in seen_keys:
                seen_keys.add(key)
                pairs.append((old, current_cnae))

        # 2) canonical cluster name (NOVA), if it differs from ANTIGA
        if current_new:
            key2 = normalize_description(current_new)
            if key2 and key2 not in seen_keys:
                seen_keys.add(key2)
                pairs.append((current_new, current_cnae))

    return pairs


async def seed_cache(database_url: str, pairs: list[tuple[str, str]]) -> int:
    conn = await asyncpg.connect(database_url, command_timeout=60)
    try:
        inserted = 0
        for desc, cnae in pairs:
            h = hash_description(desc)
            norm = normalize_description(desc)
            # Use ON CONFLICT to update only if existing metodo is lower priority
            # (golden = 90; only revisado_humano = 100 wins)
            await conn.execute(
                """
                INSERT INTO spend_classification_cache
                    (descricao_hash, descricao_normalizada, cnae, confianca, metodo, ttl)
                VALUES ($1, $2, $3, $4, 'golden', NOW() + INTERVAL '365 days')
                ON CONFLICT (descricao_hash) DO UPDATE SET
                    descricao_normalizada = EXCLUDED.descricao_normalizada,
                    cnae = EXCLUDED.cnae,
                    confianca = EXCLUDED.confianca,
                    metodo = EXCLUDED.metodo,
                    ttl = EXCLUDED.ttl
                WHERE spend_classification_cache.metodo NOT IN ('revisado_humano')
                """,
                h,
                norm,
                cnae,
                0.95,
            )
            inserted += 1
        total = await conn.fetchval("SELECT COUNT(*) FROM spend_classification_cache")
        return total
    finally:
        await conn.close()


async def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx_path", type=Path)
    parser.add_argument(
        "--sheet",
        default="Sugestão",
        help="Sheet name (default: 'Sugestão')",
    )
    args = parser.parse_args()

    if not args.xlsx_path.exists():
        print(f"ERROR: {args.xlsx_path} not found", file=sys.stderr)
        return 2

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        print("ERROR: DATABASE_URL not set", file=sys.stderr)
        return 2
    # asyncpg uses libpq form (no +asyncpg)
    dsn_pg = dsn.replace("postgresql+asyncpg://", "postgresql://")

    print(f"Reading {args.xlsx_path} sheet={args.sheet!r}...")
    pairs = extract_pairs(args.xlsx_path, args.sheet)
    print(f"Extracted {len(pairs)} (descrição, CNAE) pairs")
    if not pairs:
        return 1

    total = await seed_cache(dsn_pg, pairs)
    print(f"Cache total rows: {total}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
