"""Parse client catalog CSV/XLSX into Pydantic ParsedRow objects.

Required column: descricao_original (mapped from common Portuguese aliases like
"objeto", "descricao", "material", etc — see _HEADER_ALIASES).
Optional columns map 1:1 to spend_linhas fields; unknown columns → extras JSONB.

CSV delimiter auto-detected (`,`, `;`, or `\\t`) via csv.Sniffer.
Encoding auto-detected via chardet (utf-8 / cp1252 typically).
"""

from __future__ import annotations

import csv
import io
import unicodedata
from collections.abc import Iterator
from typing import Any

import chardet
from openpyxl import load_workbook
from pydantic import BaseModel


class CsvParseError(ValueError):
    """Raised when the catalog file is malformed."""


_KNOWN_COLUMNS = {
    "descricao_original",
    "agrupamento",
    "id_linha_origem",
    "fornecedor_atual",
    "cnpj_fornecedor",
    "valor_total",
    "quantidade",
    "uf_solicitante",
    "municipio_solicitante",
    "centro_custo",
    "data_compra",
}

# Map normalized (lowercase, no accents, stripped) Portuguese header variants
# to our internal field names. Used so pilot CSVs can be uploaded as-is.
_HEADER_ALIASES: dict[str, str] = {
    # descricao_original
    "descricao_original": "descricao_original",
    "descricao": "descricao_original",
    "descricao do material": "descricao_original",
    "descricao do produto": "descricao_original",
    "descricao do item": "descricao_original",
    "descricao item": "descricao_original",
    "descricao material": "descricao_original",
    "objeto": "descricao_original",
    "material": "descricao_original",
    "produto": "descricao_original",
    "item": "descricao_original",
    # agrupamento
    "agrupamento": "agrupamento",
    "grupo": "agrupamento",
    "grupo de material": "agrupamento",
    "grupo material": "agrupamento",
    "grup. merc": "agrupamento",
    "grup merc": "agrupamento",
    "grup. mercadoria": "agrupamento",
    "categoria": "agrupamento",
    "familia": "agrupamento",
    "classe": "agrupamento",
    # id_linha_origem
    "id_linha_origem": "id_linha_origem",
    "id": "id_linha_origem",
    "codigo": "id_linha_origem",
    "cod": "id_linha_origem",
    "codigo material": "id_linha_origem",
    "cod material": "id_linha_origem",
    # fornecedor_atual
    "fornecedor_atual": "fornecedor_atual",
    "fornecedor": "fornecedor_atual",
    "razao social": "fornecedor_atual",
    "razao social fornecedor": "fornecedor_atual",
    # cnpj_fornecedor
    "cnpj_fornecedor": "cnpj_fornecedor",
    "cnpj": "cnpj_fornecedor",
    "cnpj do fornecedor": "cnpj_fornecedor",
    # valor_total
    "valor_total": "valor_total",
    "valor": "valor_total",
    "total": "valor_total",
    "valor total": "valor_total",
    "valor (r$)": "valor_total",
    "preco": "valor_total",
    "preco total": "valor_total",
    # quantidade
    "quantidade": "quantidade",
    "qtd": "quantidade",
    "qtde": "quantidade",
    "quant": "quantidade",
    "quant.": "quantidade",
    # uf_solicitante
    "uf_solicitante": "uf_solicitante",
    "uf": "uf_solicitante",
    "estado": "uf_solicitante",
    "uf solicitante": "uf_solicitante",
    # municipio_solicitante
    "municipio_solicitante": "municipio_solicitante",
    "municipio": "municipio_solicitante",
    "cidade": "municipio_solicitante",
    # centro_custo
    "centro_custo": "centro_custo",
    "centro de custo": "centro_custo",
    "cc": "centro_custo",
    # data_compra
    "data_compra": "data_compra",
    "data": "data_compra",
    "data compra": "data_compra",
    "data da compra": "data_compra",
}


class ParsedRow(BaseModel):
    """One line of the client catalog, ready to insert into spend_linhas."""

    descricao_original: str
    agrupamento: str | None = None
    id_linha_origem: str | None = None
    fornecedor_atual: str | None = None
    cnpj_fornecedor: str | None = None
    valor_total: str | None = None
    quantidade: str | None = None
    uf_solicitante: str | None = None
    municipio_solicitante: str | None = None
    centro_custo: str | None = None
    data_compra: str | None = None
    extras: dict[str, Any] = {}


def _decode(raw: bytes) -> str:
    """Decode bytes using chardet to detect encoding (cp1252/utf-8 mostly)."""
    detection = chardet.detect(raw)
    encoding = detection.get("encoding") or "utf-8"
    try:
        return raw.decode(encoding)
    except (UnicodeDecodeError, LookupError) as exc:
        raise CsvParseError(f"unable to decode file (detected {encoding!r}): {exc}") from exc


def _is_xlsx(filename: str) -> bool:
    return filename.lower().endswith((".xlsx", ".xlsm"))


def _normalize_header(header: str) -> str:
    """Lowercase + strip + remove accents — used for alias lookup."""
    decomposed = unicodedata.normalize("NFKD", header)
    ascii_only = decomposed.encode("ascii", "ignore").decode("ascii")
    return ascii_only.lower().strip()


def _map_headers(raw_headers: list[str], overrides: dict[str, str] | None = None) -> list[str]:
    """Map each raw header to our internal field name via _HEADER_ALIASES.

    `overrides` is a user-supplied mapping (raw header → internal field name)
    that takes precedence over the alias table. Useful when the catalog has a
    non-standard column name that the user wants to map manually.

    Unknown headers keep their original name (will go to `extras`).
    """
    overrides = overrides or {}
    mapped: list[str] = []
    for h in raw_headers:
        if h is None or h == "":
            mapped.append("")
            continue
        h_str = str(h)
        if h_str in overrides:
            mapped.append(overrides[h_str])
            continue
        normalized = _normalize_header(h_str)
        mapped.append(_HEADER_ALIASES.get(normalized, h_str))
    return mapped


def _sniff_delimiter(sample: str) -> str:
    """Detect CSV delimiter. Falls back to ',' if Sniffer can't decide."""
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        return dialect.delimiter
    except csv.Error:
        return ","


def _row_from_dict(line_num: int, raw: dict[str, str]) -> ParsedRow:
    descricao = (raw.get("descricao_original") or "").strip()
    if not descricao:
        raise CsvParseError(f"empty descricao_original at line {line_num}")
    extras = {k: v for k, v in raw.items() if k and k not in _KNOWN_COLUMNS and v}
    known = {
        k: (v.strip() if isinstance(v, str) else v) or None
        for k, v in raw.items()
        if k in _KNOWN_COLUMNS
    }
    known["descricao_original"] = descricao  # already trimmed
    return ParsedRow(**known, extras=extras)


def _parse_csv_text(text: str, overrides: dict[str, str] | None = None) -> Iterator[ParsedRow]:
    sample = text[:4096]
    delimiter = _sniff_delimiter(sample)
    raw_reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    header_row = next(raw_reader, None)
    if not header_row:
        raise CsvParseError("CSV is empty")
    headers = _map_headers([h or "" for h in header_row], overrides)
    if "descricao_original" not in headers:
        raise CsvParseError(
            f"missing required column 'descricao_original' "
            f"(or known alias like 'objeto', 'descricao', 'material'). "
            f"Found columns: {header_row}"
        )
    for i, values in enumerate(raw_reader, start=2):
        # Skip rows that are entirely blank (trailing whitespace rows common in
        # Excel-saved CSVs). A row is considered "present but empty descricao"
        # only when at least one other field has content.
        if not values or all((v or "").strip() == "" for v in values):
            continue
        raw = dict(zip(headers, values, strict=False))
        yield _row_from_dict(i, raw)


def _parse_xlsx(data: bytes, overrides: dict[str, str] | None = None) -> Iterator[ParsedRow]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    raw_headers = next(rows, None)
    if not raw_headers:
        raise CsvParseError("XLSX is empty")
    headers = _map_headers([str(h) if h is not None else "" for h in raw_headers], overrides)
    if "descricao_original" not in headers:
        raise CsvParseError(
            f"missing required column 'descricao_original' "
            f"(or known alias). Found columns: {list(raw_headers)}"
        )
    for i, values in enumerate(rows, start=2):
        if not values or all(v is None or (isinstance(v, str) and v.strip() == "") for v in values):
            continue
        raw = {h: ("" if v is None else str(v)) for h, v in zip(headers, values, strict=False)}
        yield _row_from_dict(i, raw)


def parse_catalog_bytes(
    data: bytes,
    filename: str,
    overrides: dict[str, str] | None = None,
) -> Iterator[ParsedRow]:
    """Yield ParsedRow objects from a CSV or XLSX file's raw bytes.

    `overrides` maps raw header strings (as they appear in the file) to internal
    field names like "descricao_original". Use it when the file's headers are
    not in `_HEADER_ALIASES`.
    """
    if _is_xlsx(filename):
        yield from _parse_xlsx(data, overrides)
    else:
        yield from _parse_csv_text(_decode(data), overrides)


def preview_catalog_bytes(data: bytes, filename: str, sample_size: int = 5) -> dict[str, Any]:
    """Inspect a catalog file without enforcing the description column.

    Returns:
        {
            "columns": list[str],         # raw headers as they appear
            "auto_mapping": dict[str,str],# raw header → internal field if alias matched
            "sample_rows": list[list],    # first N rows (raw values, aligned to columns)
            "needs_mapping": bool,        # True if no header auto-mapped to descricao_original
        }
    """
    if _is_xlsx(filename):
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        raw_headers = next(rows_iter, None) or ()
        columns = [str(h) if h is not None else "" for h in raw_headers]
        sample: list[list[str]] = []
        for row in rows_iter:
            if len(sample) >= sample_size:
                break
            if not row or all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
                continue
            sample.append(["" if v is None else str(v) for v in row])
    else:
        text = _decode(data)
        delimiter = _sniff_delimiter(text[:4096])
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        header_row = next(reader, None) or []
        columns = [h or "" for h in header_row]
        sample = []
        for row in reader:
            if len(sample) >= sample_size:
                break
            if not row or all((v or "").strip() == "" for v in row):
                continue
            sample.append(list(row))

    auto_mapping: dict[str, str] = {}
    for col in columns:
        if not col:
            continue
        mapped = _HEADER_ALIASES.get(_normalize_header(col))
        if mapped:
            auto_mapping[col] = mapped

    needs_mapping = "descricao_original" not in auto_mapping.values()
    return {
        "columns": columns,
        "auto_mapping": auto_mapping,
        "sample_rows": sample,
        "needs_mapping": needs_mapping,
    }
