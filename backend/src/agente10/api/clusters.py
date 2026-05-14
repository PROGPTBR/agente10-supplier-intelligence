"""REST endpoints for cluster review and shortlist viewing."""

from __future__ import annotations

import io
from datetime import date as date_t
from datetime import datetime
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel
from sqlalchemy import text

from agente10.api.uploads import get_tenant_id
from agente10.cache import classification_cache as cache
from agente10.core.db import get_session_factory
from agente10.core.tenancy import tenant_context
from agente10.worker.client import get_pool

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

router = APIRouter(prefix="/api/v1", tags=["clusters"])


class ClusterSummary(BaseModel):
    id: UUID
    nome_cluster: str
    nome_cluster_refinado: str | None
    cnae: str | None
    cnae_descricao: str | None
    cnae_confianca: float | None
    cnae_metodo: str | None
    cnaes_secundarios: list[str]
    num_linhas: int
    revisado_humano: bool
    shortlist_size: int


class ClusterDetail(BaseModel):
    id: UUID
    upload_id: UUID
    nome_cluster: str
    nome_cluster_refinado: str | None
    cnae: str | None
    cnae_descricao: str | None
    cnae_confianca: float | None
    cnae_metodo: str | None
    cnaes_secundarios: list[str]
    num_linhas: int
    revisado_humano: bool
    notas_revisor: str | None
    shortlist_gerada: bool
    sample_linhas: list[str]


class ShortlistEntryView(BaseModel):
    """One supplier row in the shortlist — grouped by cnpj_basico (8-digit company root).

    `cnpj` is the representative filial (matriz when present, else highest capital);
    `filiais_count` is the total filiais of this company present in our `empresas`
    table (limited by the pilot trim — see project memory).
    """

    cnpj_basico: str
    cnpj: str
    razao_social: str
    nome_fantasia: str | None
    capital_social: float | None
    uf: str | None
    municipio: str | None
    data_abertura: date_t | None
    rank_estagio3: int
    filiais_count: int


class FilialView(BaseModel):
    cnpj: str
    razao_social: str
    nome_fantasia: str | None
    capital_social: float | None
    uf: str | None
    municipio: str | None
    cep: str | None
    endereco: str | None
    data_abertura: date_t | None
    situacao_cadastral: str | None
    is_matriz: bool


class ClusterPatch(BaseModel):
    cnae: str | None = None
    cnaes_secundarios: list[str] | None = None
    notas_revisor: str | None = None
    revisado_humano: bool | None = None
    handoff_rfx: bool | None = None


@router.get("/uploads/{upload_id}/clusters", response_model=list[ClusterSummary])
async def list_clusters_for_upload(
    upload_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
    metodo: str | None = Query(default=None),
    revisado: bool | None = Query(default=None),
) -> list[ClusterSummary]:
    factory = get_session_factory()
    sql = """
        SELECT
            c.id, c.nome_cluster, c.nome_cluster_refinado,
            c.cnae, ct.denominacao AS cnae_descricao,
            c.cnae_confianca, c.cnae_metodo, c.cnaes_secundarios,
            c.num_linhas, c.revisado_humano,
            (SELECT COUNT(*) FROM supplier_shortlists s WHERE s.cnae = c.cnae
                AND s.tenant_id = c.tenant_id) AS shortlist_size
        FROM spend_clusters c
        LEFT JOIN cnae_taxonomy ct ON ct.codigo = c.cnae
        WHERE c.upload_id = :u
    """
    params: dict[str, object] = {"u": str(upload_id)}
    if metodo:
        sql += " AND c.cnae_metodo = :m"
        params["m"] = metodo
    if revisado is not None:
        sql += " AND c.revisado_humano = :r"
        params["r"] = revisado
    sql += " ORDER BY c.nome_cluster"

    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            rows = (await session.execute(text(sql), params)).all()
    return [
        ClusterSummary(
            id=r.id,
            nome_cluster=r.nome_cluster,
            nome_cluster_refinado=r.nome_cluster_refinado,
            cnae=r.cnae,
            cnae_descricao=r.cnae_descricao,
            cnae_confianca=r.cnae_confianca,
            cnae_metodo=r.cnae_metodo,
            cnaes_secundarios=list(r.cnaes_secundarios or []),
            num_linhas=r.num_linhas,
            revisado_humano=r.revisado_humano,
            shortlist_size=int(r.shortlist_size),
        )
        for r in rows
    ]


@router.get("/clusters/{cluster_id}", response_model=ClusterDetail)
async def get_cluster(
    cluster_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
) -> ClusterDetail:
    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            row = (
                await session.execute(
                    text("""
                        SELECT c.id, c.upload_id, c.nome_cluster,
                               c.nome_cluster_refinado, c.cnae,
                               ct.denominacao AS cnae_descricao,
                               c.cnae_confianca, c.cnae_metodo,
                               c.cnaes_secundarios,
                               c.num_linhas,
                               c.revisado_humano, c.notas_revisor,
                               c.shortlist_gerada
                        FROM spend_clusters c
                        LEFT JOIN cnae_taxonomy ct ON ct.codigo = c.cnae
                        WHERE c.id = :i
                        """),
                    {"i": str(cluster_id)},
                )
            ).first()
            if not row:
                raise HTTPException(404, "cluster not found")
            sample_rows = (
                await session.execute(
                    text(
                        "SELECT descricao_original FROM spend_linhas "
                        "WHERE cluster_id = :i ORDER BY id LIMIT 5"
                    ),
                    {"i": str(cluster_id)},
                )
            ).all()
    return ClusterDetail(
        id=row.id,
        upload_id=row.upload_id,
        nome_cluster=row.nome_cluster,
        nome_cluster_refinado=row.nome_cluster_refinado,
        cnae=row.cnae,
        cnae_descricao=row.cnae_descricao,
        cnae_confianca=row.cnae_confianca,
        cnae_metodo=row.cnae_metodo,
        cnaes_secundarios=list(row.cnaes_secundarios or []),
        num_linhas=row.num_linhas,
        revisado_humano=row.revisado_humano,
        notas_revisor=row.notas_revisor,
        shortlist_gerada=row.shortlist_gerada,
        sample_linhas=[s.descricao_original for s in sample_rows],
    )


@router.get("/clusters/{cluster_id}/shortlist", response_model=list[ShortlistEntryView])
async def get_cluster_shortlist(
    cluster_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
    uf: str | None = Query(default=None, max_length=2),
    municipio: str | None = Query(default=None),
) -> list[ShortlistEntryView]:
    """Top suppliers for a cluster, grouped by cnpj_basico (one row per company).

    Without filters: source is supplier_shortlists (the curator's ranked top-10).
    With UF/município filters: re-queries empresas table directly so the result
    set can extend beyond the precomputed shortlist (otherwise filtering would
    often produce empty tables).
    """
    from agente10.config.shortlist import parse_from_metadados

    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            cnae_row = (
                await session.execute(
                    text(
                        "SELECT c.cnae, c.cnaes_secundarios, u.metadados "
                        "FROM spend_clusters c "
                        "JOIN spend_uploads u ON u.id = c.upload_id "
                        "WHERE c.id = :i"
                    ),
                    {"i": str(cluster_id)},
                )
            ).first()
            if not cnae_row:
                raise HTTPException(404, "cluster not found")
            if not cnae_row.cnae:
                return []
            cnaes_alvo = [cnae_row.cnae] + list(cnae_row.cnaes_secundarios or [])
            view_size = parse_from_metadados(cnae_row.metadados).size

            if uf or municipio:
                # Filtered: query empresas directly, group by cnpj_basico
                sql = """
                    WITH per_company AS (
                        SELECT DISTINCT ON (substring(e.cnpj, 1, 8))
                            substring(e.cnpj, 1, 8) AS cnpj_basico,
                            e.cnpj, e.razao_social, e.nome_fantasia,
                            e.capital_social, e.uf, e.municipio, e.data_abertura
                        FROM empresas e
                        WHERE (e.cnae_primario = ANY(:cnaes)
                               OR e.cnaes_secundarios && :cnaes)
                          AND e.situacao_cadastral = 'ATIVA'
                          AND (CAST(:uf AS text) IS NULL OR e.uf = :uf)
                          AND (CAST(:m AS text) IS NULL OR e.municipio = :m)
                        ORDER BY substring(e.cnpj, 1, 8),
                                 CASE WHEN substring(e.cnpj, 9, 4) = '0001' THEN 0 ELSE 1 END,
                                 e.capital_social DESC NULLS LAST,
                                 e.data_abertura ASC NULLS LAST
                    ),
                    ranked AS (
                        SELECT
                            pc.*,
                            ROW_NUMBER() OVER (
                                ORDER BY pc.capital_social DESC NULLS LAST,
                                         pc.data_abertura ASC NULLS LAST
                            ) AS rank
                        FROM per_company pc
                        LIMIT :lim
                    )
                    SELECT r.*,
                           (SELECT COUNT(*) FROM empresas e2
                            WHERE substring(e2.cnpj, 1, 8) = r.cnpj_basico) AS filiais_count
                    FROM ranked r
                    ORDER BY r.rank
                """
                params: dict[str, object] = {
                    "cnaes": cnaes_alvo,
                    "uf": uf,
                    "m": municipio,
                    "lim": view_size,
                }
            else:
                # Unfiltered: use precomputed supplier_shortlists, then group
                sql = """
                    WITH base AS (
                        SELECT DISTINCT ON (s.cnpj_fornecedor)
                            s.cnpj_fornecedor AS cnpj, s.rank_estagio3
                        FROM supplier_shortlists s
                        WHERE s.cnae = ANY(:cnaes) AND s.tenant_id = :t
                        ORDER BY s.cnpj_fornecedor, s.rank_estagio3
                    ),
                    per_company AS (
                        SELECT DISTINCT ON (substring(e.cnpj, 1, 8))
                            substring(e.cnpj, 1, 8) AS cnpj_basico,
                            e.cnpj, e.razao_social, e.nome_fantasia,
                            e.capital_social, e.uf, e.municipio, e.data_abertura,
                            b.rank_estagio3
                        FROM base b
                        JOIN empresas e ON substring(e.cnpj, 1, 8) = substring(b.cnpj, 1, 8)
                        ORDER BY substring(e.cnpj, 1, 8),
                                 CASE WHEN substring(e.cnpj, 9, 4) = '0001' THEN 0 ELSE 1 END,
                                 e.capital_social DESC NULLS LAST,
                                 e.data_abertura ASC NULLS LAST
                    )
                    SELECT pc.*,
                           (SELECT COUNT(*) FROM empresas e2
                            WHERE substring(e2.cnpj, 1, 8) = pc.cnpj_basico) AS filiais_count,
                           ROW_NUMBER() OVER (ORDER BY pc.rank_estagio3) AS rank
                    FROM per_company pc
                    ORDER BY pc.rank_estagio3
                    LIMIT :lim
                """
                params = {"cnaes": cnaes_alvo, "t": str(tenant_id), "lim": view_size}

            rows = (await session.execute(text(sql), params)).all()

    return [
        ShortlistEntryView(
            cnpj_basico=r.cnpj_basico,
            cnpj=r.cnpj,
            razao_social=r.razao_social,
            nome_fantasia=r.nome_fantasia,
            capital_social=float(r.capital_social) if r.capital_social is not None else None,
            uf=r.uf,
            municipio=r.municipio,
            data_abertura=r.data_abertura,
            rank_estagio3=int(r.rank if uf or municipio else r.rank_estagio3),
            filiais_count=int(r.filiais_count),
        )
        for r in rows
    ]


@router.get(
    "/clusters/{cluster_id}/empresa/{cnpj_basico}/filiais",
    response_model=list[FilialView],
)
async def get_company_filiais(
    cluster_id: UUID,
    cnpj_basico: str,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
) -> list[FilialView]:
    """All filiais of a company (same first 8 CNPJ digits) present in our empresas
    table. Matriz (filial code 0001) is returned first.

    Note: the pilot's empresas trim only loaded subsets of CNAEs, so this may
    return fewer filiais than reality. The trim is documented in project memory.
    """
    if len(cnpj_basico) != 8 or not cnpj_basico.isdigit():
        raise HTTPException(400, "cnpj_basico must be 8 digits")
    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            # Validate cluster exists (RLS enforced)
            exists = (
                await session.execute(
                    text("SELECT 1 FROM spend_clusters WHERE id = :i"),
                    {"i": str(cluster_id)},
                )
            ).first()
            if not exists:
                raise HTTPException(404, "cluster not found")
            rows = (
                await session.execute(
                    text("""
                        SELECT cnpj, razao_social, nome_fantasia, capital_social,
                               uf, municipio, cep, endereco, data_abertura,
                               situacao_cadastral,
                               substring(cnpj, 9, 4) = '0001' AS is_matriz
                        FROM empresas
                        WHERE substring(cnpj, 1, 8) = :b
                        ORDER BY is_matriz DESC, capital_social DESC NULLS LAST, cnpj
                    """),
                    {"b": cnpj_basico},
                )
            ).all()
    return [
        FilialView(
            cnpj=r.cnpj,
            razao_social=r.razao_social,
            nome_fantasia=r.nome_fantasia,
            capital_social=float(r.capital_social) if r.capital_social is not None else None,
            uf=r.uf,
            municipio=r.municipio,
            cep=r.cep,
            endereco=r.endereco,
            data_abertura=r.data_abertura,
            situacao_cadastral=r.situacao_cadastral,
            is_matriz=bool(r.is_matriz),
        )
        for r in rows
    ]


@router.patch("/clusters/{cluster_id}", response_model=ClusterDetail)
async def patch_cluster(
    cluster_id: UUID,
    body: ClusterPatch,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
) -> ClusterDetail:
    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            current = (
                await session.execute(
                    text(
                        "SELECT nome_cluster, cnae, cnae_metodo, cnaes_secundarios "
                        "FROM spend_clusters WHERE id = :i"
                    ),
                    {"i": str(cluster_id)},
                )
            ).first()
            if not current:
                raise HTTPException(404, "cluster not found")
            cnae_changed = body.cnae is not None and body.cnae != current.cnae
            current_sec = sorted(set(current.cnaes_secundarios or []))
            new_sec = (
                sorted(set(body.cnaes_secundarios))
                if body.cnaes_secundarios is not None
                else current_sec
            )
            secundarios_changed = body.cnaes_secundarios is not None and new_sec != current_sec

            codes_to_validate: set[str] = set()
            if cnae_changed:
                codes_to_validate.add(body.cnae)  # type: ignore[arg-type]
            if secundarios_changed:
                codes_to_validate.update(new_sec)
            for code in codes_to_validate:
                taxonomy_hit = (
                    await session.execute(
                        text("SELECT denominacao FROM cnae_taxonomy WHERE codigo = :c"),
                        {"c": code},
                    )
                ).first()
                if not taxonomy_hit:
                    raise HTTPException(422, f"cnae '{code}' not found in cnae_taxonomy")

            if cnae_changed:
                # Audit trail BEFORE the UPDATE — preserves the human-correction signal
                # as future training data even if the cluster gets reclassified later.
                await session.execute(
                    text("""
                        INSERT INTO cnae_correction_audit (
                            tenant_id, cluster_id, descricao_cluster,
                            cnae_antes, metodo_antes, cnae_depois, notas_revisor
                        ) VALUES (:t, :i, :d, :ca, :ma, :cd, :n)
                    """),
                    {
                        "t": str(tenant_id),
                        "i": str(cluster_id),
                        "d": current.nome_cluster,
                        "ca": current.cnae,
                        "ma": current.cnae_metodo,
                        "cd": body.cnae,
                        "n": body.notas_revisor,
                    },
                )

            updates: list[str] = []
            params: dict[str, object] = {"i": str(cluster_id)}
            if body.cnae is not None:
                updates.append("cnae = :cnae")
                params["cnae"] = body.cnae
                updates.append("cnae_metodo = 'revisado_humano'")
            if secundarios_changed:
                updates.append("cnaes_secundarios = :sec")
                params["sec"] = new_sec
            if body.notas_revisor is not None:
                updates.append("notas_revisor = :n")
                params["n"] = body.notas_revisor
            if body.revisado_humano is not None:
                updates.append("revisado_humano = :r")
                params["r"] = body.revisado_humano
            if cnae_changed or secundarios_changed:
                updates.append("shortlist_gerada = false")
            if updates:
                await session.execute(
                    text(f"UPDATE spend_clusters SET {', '.join(updates)} WHERE id = :i"),
                    params,
                )

            if cnae_changed:
                # Cache reflects the human-validated CNAE for this description.
                # Priority 'revisado_humano' won't be downgraded by future curator runs.
                await cache.upsert(
                    session,
                    current.nome_cluster,
                    body.cnae,  # type: ignore[arg-type]
                    confianca=1.0,
                    metodo="revisado_humano",
                )

    if cnae_changed or secundarios_changed:
        pool = get_pool()
        await pool.enqueue_job(
            "run_regenerate_shortlist",
            str(cluster_id),
            str(tenant_id),
        )

    return await get_cluster(cluster_id, tenant_id=tenant_id)


_EXPORT_COLUMNS = [
    "Cluster",
    "CNAE",
    "CNAE descrição",
    "Rank",
    "Razão social",
    "Nome fantasia",
    "CNPJ (matriz)",
    "Filiais",
    "Capital social (BRL)",
    "UF",
    "Município",
    "Data abertura",
]


def _shortlist_query(*, single_cluster: bool) -> str:
    """Top-10 shortlist rows per cluster joined with empresas.

    `single_cluster=True` adds a cluster-id filter; otherwise filters by upload.
    """
    where = "c.id = :cluster_id" if single_cluster else "c.upload_id = :upload_id"
    return f"""
        WITH target_clusters AS (
            SELECT c.id, c.nome_cluster, c.nome_cluster_refinado, c.cnae,
                   ARRAY[c.cnae] || COALESCE(c.cnaes_secundarios, ARRAY[]::varchar[]) AS cnaes_alvo
            FROM spend_clusters c
            WHERE {where} AND c.cnae IS NOT NULL
        ),
        base AS (
            SELECT DISTINCT ON (tc.id, s.cnpj_fornecedor)
                tc.id AS cluster_id,
                COALESCE(tc.nome_cluster_refinado, tc.nome_cluster) AS cluster_label,
                tc.cnae AS cluster_cnae,
                s.cnpj_fornecedor AS cnpj,
                s.rank_estagio3
            FROM target_clusters tc
            JOIN supplier_shortlists s
              ON s.cnae = ANY(tc.cnaes_alvo) AND s.tenant_id = :tenant_id
            ORDER BY tc.id, s.cnpj_fornecedor, s.rank_estagio3
        ),
        per_company AS (
            SELECT DISTINCT ON (b.cluster_id, substring(e.cnpj, 1, 8))
                b.cluster_id, b.cluster_label, b.cluster_cnae,
                substring(e.cnpj, 1, 8) AS cnpj_basico,
                e.cnpj, e.razao_social, e.nome_fantasia,
                e.capital_social, e.uf, e.municipio, e.data_abertura,
                b.rank_estagio3
            FROM base b
            JOIN empresas e ON substring(e.cnpj, 1, 8) = substring(b.cnpj, 1, 8)
            ORDER BY b.cluster_id, substring(e.cnpj, 1, 8),
                     CASE WHEN substring(e.cnpj, 9, 4) = '0001' THEN 0 ELSE 1 END,
                     e.capital_social DESC NULLS LAST,
                     e.data_abertura ASC NULLS LAST
        )
        SELECT pc.*, ct.denominacao AS cnae_descricao,
               (SELECT COUNT(*) FROM empresas e2
                WHERE substring(e2.cnpj, 1, 8) = pc.cnpj_basico) AS filiais_count,
               ROW_NUMBER() OVER (
                   PARTITION BY pc.cluster_id ORDER BY pc.rank_estagio3
               ) AS rank
        FROM per_company pc
        LEFT JOIN cnae_taxonomy ct ON ct.codigo = pc.cluster_cnae
        ORDER BY pc.cluster_label, rank
    """


def _build_xlsx(rows: list[dict[str, object]], sheet_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31] or "Shortlist"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for col_idx, name in enumerate(_EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
    for r_idx, row in enumerate(rows, start=2):
        ws.cell(row=r_idx, column=1, value=row["cluster_label"])
        ws.cell(row=r_idx, column=2, value=row["cluster_cnae"])
        ws.cell(row=r_idx, column=3, value=row["cnae_descricao"])
        ws.cell(row=r_idx, column=4, value=row["rank"])
        ws.cell(row=r_idx, column=5, value=row["razao_social"])
        ws.cell(row=r_idx, column=6, value=row["nome_fantasia"])
        ws.cell(row=r_idx, column=7, value=row["cnpj"])
        ws.cell(row=r_idx, column=8, value=row["filiais_count"])
        ws.cell(row=r_idx, column=9, value=row["capital_social"])
        ws.cell(row=r_idx, column=10, value=row["uf"])
        ws.cell(row=r_idx, column=11, value=row["municipio"])
        ws.cell(row=r_idx, column=12, value=row["data_abertura"])
    widths = [38, 10, 50, 6, 50, 50, 18, 8, 18, 5, 24, 12]
    for col_idx, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _rows_to_dicts(rows) -> list[dict[str, object]]:
    return [
        {
            "cluster_label": r.cluster_label,
            "cluster_cnae": r.cluster_cnae,
            "cnae_descricao": r.cnae_descricao,
            "rank": int(r.rank),
            "razao_social": r.razao_social,
            "nome_fantasia": r.nome_fantasia,
            "cnpj": r.cnpj,
            "filiais_count": int(r.filiais_count),
            "capital_social": (float(r.capital_social) if r.capital_social is not None else None),
            "uf": r.uf,
            "municipio": r.municipio,
            "data_abertura": r.data_abertura,
        }
        for r in rows
    ]


@router.get("/uploads/{upload_id}/shortlist.xlsx")
async def export_upload_shortlist(
    upload_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
) -> StreamingResponse:
    """Single XLSX with the top-10 shortlist of every classified cluster in an upload."""
    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            meta = (
                await session.execute(
                    text("SELECT nome_arquivo FROM spend_uploads WHERE id = :u"),
                    {"u": str(upload_id)},
                )
            ).first()
            if not meta:
                raise HTTPException(404, "upload not found")
            rows = (
                await session.execute(
                    text(_shortlist_query(single_cluster=False)),
                    {"upload_id": str(upload_id), "tenant_id": str(tenant_id)},
                )
            ).all()
    xlsx_bytes = _build_xlsx(_rows_to_dicts(rows), sheet_title="Shortlist")
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    safe_name = "".join(c if c.isalnum() or c in "-_." else "_" for c in meta.nome_arquivo)[:40]
    filename = f"shortlist-{safe_name}-{stamp}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/clusters/{cluster_id}/shortlist.xlsx")
async def export_cluster_shortlist(
    cluster_id: UUID,
    tenant_id: UUID = Depends(get_tenant_id),  # noqa: B008
) -> StreamingResponse:
    """XLSX with the shortlist of a single cluster."""
    factory = get_session_factory()
    async with factory() as session, session.begin():
        async with tenant_context(session, tenant_id):
            meta = (
                await session.execute(
                    text(
                        "SELECT COALESCE(nome_cluster_refinado, nome_cluster) AS label "
                        "FROM spend_clusters WHERE id = :i"
                    ),
                    {"i": str(cluster_id)},
                )
            ).first()
            if not meta:
                raise HTTPException(404, "cluster not found")
            rows = (
                await session.execute(
                    text(_shortlist_query(single_cluster=True)),
                    {"cluster_id": str(cluster_id), "tenant_id": str(tenant_id)},
                )
            ).all()
    xlsx_bytes = _build_xlsx(_rows_to_dicts(rows), sheet_title=meta.label)
    stamp = datetime.now().strftime("%Y%m%d-%H%M")
    safe_name = "".join(c if c.isalnum() or c in "-_." else "_" for c in meta.label)[:40]
    filename = f"shortlist-{safe_name}-{stamp}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type=XLSX_MEDIA_TYPE,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
